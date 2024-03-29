# 전일 종가 정보 추출 및 매수가 대비 수익률, 수익금액 계산하여 엑셀파일에 첨부

# 종가 정보 크롤링 -> 네이버 검색이용
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import pandas as pd
import openpyxl

wd = webdriver.Chrome()
wd.implicitly_wait(3)  # 3초간 반복적으로 작업이 수행될 때까지 반복하는 것
url = r'https://finance.naver.com/'
wd.get(url)
time.sleep(1)

exprice_lst = []  # 전일 종가 정보 저장(오전 장시작 전 기준: 현재가->전일 종가)

stock_data = list(pd.read_excel('user_infos.xlsx', engine='openpyxl', header=0, index_col=0)['종목명'][:])
# stock_data = ['두산에너빌리티', '카카오', '삼성전자', 'lg에너지솔루션']
for i in stock_data:
    search = wd.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div[2]/form/fieldset/div/div[1]/input')
    search.send_keys(f"{i}")  # 보유 중인 주식명 기입
    time.sleep(1)
    wd.find_element(By.XPATH, '/html/body/div[2]/div[1]/div/div[2]/form/fieldset/div/div[2]/div/div[1]/div/div/ul/li/a').click()
    time.sleep(1)
    exprice = wd.find_elements(By.XPATH, '/html/body/div[3]/div[2]/div[2]/div[1]/div[1]/div[1]/div/p[1]/em/span')  # 전일 종가 정보
    for j in range(len(exprice)):
        exprice[j] = exprice[j].text
    exprice_lst.append(int(''.join(exprice).replace(',', "")))
print(exprice_lst)

fpath = 'user_infos.xlsx'  # user_info 파일에 새로운 정보 추가(전일종가, 평가금액, 평가손익, 손익률)
wb = openpyxl.load_workbook(fpath)
ws = wb['user1']
ws['F1'] = '전일 종가'
ws['G1'] = '평가 금액'
ws['H1'] = '평가 손익'
ws['I1'] = '손익률'
for k in range(2, len(stock_data)+2):
    ws[f'F{k}'] = exprice_lst[k-2]
    ws[f'G{k}'] = int(ws[f'F{k}'].value) * int(ws[f'D{k}'].value)  # 평가금액 = 전일종가 * 보유수량
    ws[f'H{k}'] = int(ws[f'G{k}'].value) - int(ws[f'E{k}'].value)  # 평가손익 = 평가금액 - 매수금액
    if int(ws[f'E{k}'].value) != 0:
        ws[f'I{k}'] = int(ws[f'H{k}'].value) / int(ws[f'E{k}'].value) * 100  # 손익률 = 평가손익 / 매수금액
    else:
        ws[f'I{k}'] = 0
wb.save(fpath)
