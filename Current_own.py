'''현재 보유중인 주식과 매수 평단가, 추가로 관심있는 종목들에 대한 정보를 사용자로부터 받음'''
import xlsxwriter
import openpyxl
'''
def register_stock_info():
    own = int(input('보유중인 종목 개수를 입력해주세요: '))
    for n in range(own):
        i = list(input(f"보유 중인 종목 [주식명, 매수 평단가]를 입력해 주세요{{{n+1}}}: ").split())
        stock_info[i[0]] = int(i[1])

stock_info = {}
register_stock_info()
print(stock_info)
'''
# sing up 할 때 사용, 수정할 경우 모두 초기화 되기 때문에
def register_stock(stock):
    for i in range(n):
        stock.append(input(f"보유 중인 종목의 주식명을 입력해 주세요{{{i+1}}}: "))
    fin = input("수정이 필요 합니까? (Y/N): ")
    if fin == 'Y' or fin == 'y':
        stock = []   # 모두 초기화 하고 재입력
        register_stock(stock)
# stock = ['삼성전자', '두산에너빌리티', '카카오', '코리안리']
def register_buy_price(buy_info):
    for i in range(n):
        buy_info.append(list(map(int,input(f"{stock[i]}의 매수 평균가, 수량을 입력해 주세요{{{i+1}}}: ").split(','))))
    fin2 = input("수정이 필요하십니까? (Y/N): ")
    if fin2 == 'Y' or fin2 == 'y':
        buy_info = []   # 모두 초기화 하고 재입력
        register_buy_price(buy_info)
# buy_info = [[71000, 16], [20050, 99], [122000, 8], [9980, 100]]
def register_stockplus(stockplus):
    while True:
        pluslist = list((input("관심 종목을 추가해 주세요: ").split(" ")))
        fin3 = input("수정이 필요하십니까? (Y/N): ")
        if fin3 == 'Y' or fin3 == 'y':
            continue
        break
    for i in pluslist:
        stockplus.append(i)
# stockplus = ['네이버', '포스코홀딩스']
stock = []
buy_info = []
stockplus = []
buy_lst = []
n = int(input('보유 중인 종목 개수를 입력해주세요: '))

register_stock(stock)
register_buy_price(buy_info)
register_stockplus(stockplus)
m = len(stockplus)

print(stock)
print(buy_info)
print(stockplus)

'''차후 추가사항
1. 입력받은 데이터를 엑셀파일로 저장
2. 엑셀형식 정하기 -> column: 주식 종목명 / 매수평단가
3. 관심종목은 주식종목명 하단에 추가되며 매수평단가는 0으로 설정'''

'''
wb = openpyxl.Workbook()
ws = wb.create_sheet('user1')

ws['A1'] = '구분'
ws['B1'] = '종목명'
ws['C1'] = '매수평단가'

wb.save('user_infos.xlsx')
'''
wb = xlsxwriter.Workbook('user_infos.xlsx')
ws = wb.add_worksheet('user1')
ws.write('A1', "구분")
ws.write('B1', "종목명")
ws.write('C1', "평균매입가")
ws.write('D1', "잔고수량")
ws.write('E1', "매수금액")
# ws.write('F1', "평가손익")   # 추후 계산 > 현재가 확인 파일 생성 후
# ws.write('G1', "수익률")    # 추후 계산 > 현재가 확인 파일 생성 후

for i in range(n+m):
    if i < n:
        buy_lst.append([stock[i],buy_info[i][0], buy_info[i][1]])  # [['삼성전자',71000, 16], ...]
    else:
        buy_lst.append([stockplus[i-n], 0, 0])  # [['삼성전자',71000, 16], ...]

for i in range(n+m):
    ws.write(f'A{i+2}', i+1)  # 번호
    ws.write(f'B{i+2}', buy_lst[i][0])  # 종목명
    if i < n:  # 관심종목 전까지만 추가 정보저장
        ws.write(f'C{i+2}', buy_lst[i][1])  # 평균매입가
        ws.write(f'D{i+2}', buy_lst[i][2])  # 잔고수량
        ws.write(f'E{i+2}', buy_lst[i][1]*buy_lst[i-2][2])  # 매수금액 = 평균매입가*잔고수량
        #    ws.write(f'F{i}',f'=E{i} - {buy_price}')
        #    ws.write(f'G{i}',f'=F{i}/{buy_price}*100')
    else:
        ws.write(f'C{i + 2}', 0)  # 평균매입가
        ws.write(f'D{i + 2}', 0)  # 잔고수량
        ws.write(f'E{i + 2}', 0)
wb.close()